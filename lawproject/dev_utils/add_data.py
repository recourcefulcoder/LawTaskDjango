import os
import sys
from pathlib import Path

import django

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'lawproject.settings')
sys.path.append(f"{BASE_DIR}")
django.setup()

from reference.models import Category


def json_add(data: list, parent: Category | None = None) -> bool:
    """
    :param data: list of categories to be added; each category in this list is
    a dictionary; each category must have 'children' parameter, which defines
    what children of this category are
    :param parent: parent of all categories in the list
    :return: returns whether any of this data's categories require update
    """
    ret_bool = False
    for elem in data:
        el_data = elem.copy()
        del el_data["children"]
        if parent is not None:
            el_data["parent"] = parent
        new_node = Category.objects.create(
            **el_data
        )
        ret_bool = new_node.auto_updates or \
            json_add(elem["children"], parent=new_node)
        new_node.auto_updates = ret_bool or new_node.auto_updates
        new_node.save()
    return ret_bool


def find_parent(data: list, name: str) -> int:
    cnt = 0
    length = len(data)
    while cnt < length and data[cnt]["name"] != name:
        cnt += 1
    if cnt == length:
        return -1
    return cnt


def find_nested_parent(data: list, name: str):
    cnt1 = 0
    length = len(data)
    while cnt1 < length:
        cnt2 = 0
        nested = data[cnt1]["children"]
        len2 = len(nested)
        while cnt2 < len2 and nested[cnt2]["name"] != name:
            cnt2 += 1
        if cnt2 < len2 and nested[cnt2]["name"] == name:
            return cnt1, cnt2
        cnt1 += 1
    return -1, -1


def excel_add(filename: str):
    wb = load_workbook(filename)
    sh1 = wb["Категория1"]
    sh2 = wb["Категория2"]
    sh3 = wb["Категория3"]
    data = []
    for row in sh1.iter_rows(min_row=2):
        data.append({
            "name": row[0].value,
            "block": row[1].value,
            "auto_updates": row[2].value is not None,
            "type_id": row[3].value,
            "children": [],
        })
    for row in sh2.iter_rows(min_row=2):
        if row[0].value is None:
            break
        new_block = {
            "name": row[1].value,
            "type_id": row[2].value,
            "block": row[3].value,
            "auto_updates": row[4].value is not None,
            "children": []
        }
        ind = find_parent(data, row[0].value)
        if ind == -1:
            continue
        data[ind]["children"].append(
            new_block
        )
    for row in sh3.iter_rows(min_row=2):
        if row[0].value is None:
            break
        ind1, ind2 = find_nested_parent(data, row[0].value)
        if ind1 == -1:
            continue
        new_block = {
            "name": row[1].value,
            "type_id": row[2].value,
            "block": row[3].value,
            "auto_updates": row[4].value is not None,
            "children": []
        }
        data[ind1]["children"][ind2]["children"].append(
            new_block
        )

    json_add(data)


if __name__ == "__main__":
    # print(Category.objects.all())
    Category.objects.all().delete()
    excel_add("data.xlsx")
